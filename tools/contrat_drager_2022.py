#  Copyright (c) 2020 Brice Nord, Romuald Kliglich, Alexandre Jaborska, Philomène Mazand.
#  This file is part of the BiomAid distribution.
#  This program is free software: you can redistribute it and/or modify
#  it under the terms of the GNU General Public License as published by
#  the Free Software Foundation, version 3.
#  This program is distributed in the hope that it will be useful, but
#  WITHOUT ANY WARRANTY; without even the implied warranty of
#  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the GNU
#  General Public License for more details.
#  You should have received a copy of the GNU General Public License
#  along with this program. If not, see <http://www.gnu.org/licenses/>.
import os.path

import openpyxl

from assetplusconnect.models import BEq1996, HistoEq
from pandas import read_excel

CONTRAT_CODE = '190350'
FILENAME = '/home/alexandre/tmp/draeger/ECHEANCIER CH AMIENS doc drager.xlsx'
N_SHEETS = 8


def run():
    print("Test contrat Dräger")

    all_eqpts_list = list(BEq1996.objects.using('gmao').filter(marque='DRAGER').values())
    all_eqpt_dict = {eqpt['n_imma']: eqpt for eqpt in all_eqpts_list}
    all_eqpt_ns_dict = {eqpt['n_seri'].upper().replace(' ', '').replace('-', ''): eqpt for eqpt in all_eqpts_list}
    print("Nombre total d'équipements Dräger :", len(all_eqpt_dict))

    contrat_eqpts = list(HistoEq.objects.using('gmao').filter(n_contrat=CONTRAT_CODE).values())
    print("Nombre d'équipements Dräger dans le contrat (toutes années) :", len(contrat_eqpts))

    contrat_lnks = {}
    for lnk in contrat_eqpts:
        data = (lnk['annee_exo'], lnk['ttc_annee'])
        if lnk['n_imma'] not in contrat_lnks.keys():
            contrat_lnks[lnk['n_imma']] = [data]
        else:
            contrat_lnks[lnk['n_imma']].append(data)

    all_serials = set()

    for sheet in range(N_SHEETS):
        df = read_excel(
            FILENAME, sheet_name=sheet, skiprows=2, header=None, names=['code', 'name', 'qte', 'cond', 'amount', 'model', 'serial']
        )
        all_serials = all_serials.union(set(df['serial']))
        # print(sheet, df)

    found = 0
    not_found = 0
    sn_imma = {}
    for serial in all_serials:
        serial_c = str(serial).upper().replace(' ', '').replace('-', '')
        done = False
        for r_serial in all_eqpt_ns_dict.keys():
            if serial_c in r_serial:
                print(serial, '==>', all_eqpt_ns_dict[r_serial]['n_imma'])
                sn_imma[str(serial)] = all_eqpt_ns_dict[r_serial]['n_imma']
                found += 1
                done = True
                break
        if not done:
            print(serial, '==X')
            not_found += 1
    print("Total", len(all_serials))
    print("Found", found, len(sn_imma))
    print("Not found", not_found)
    # print(sn_imma)

    wb = openpyxl.load_workbook(FILENAME)
    for ws in wb.worksheets:
        row_idx = 3
        val = 0
        ws.cell(2, 8).value = "Inventaire"
        ws.cell(2, 9).value = "N° Série"
        ws.cell(2, 10).value = "Contrat"
        ws.cell(2, 11).value = "Fin garantie"
        ws.cell(2, 12).value = "Réforme"
        while val != 'None':
            val = str(ws.cell(row_idx, 7).value)
            if val and val in sn_imma:
                ws.cell(row_idx, 8).value = sn_imma[val]
                ws.cell(row_idx, 9).value = all_eqpt_dict[sn_imma[val]]['n_seri']
                ws.cell(row_idx, 10).value = ', '.join(
                    str(y[0]) + ': {:2.2f} €'.format(float(y[1])) for y in contrat_lnks[sn_imma[val]]
                )
                ws.cell(row_idx, 11).value = all_eqpt_dict[sn_imma[val]]['fdg']
                ws.cell(row_idx, 12).value = all_eqpt_dict[sn_imma[val]]['date_refor']
            row_idx += 1
    ws = wb.create_sheet("Non trouvé")
    row_idx = 2
    ws.cell(1, 1).value = "Inventaire"
    ws.cell(1, 2).value = "N° Série"
    ws.cell(1, 3).value = "Contrat"
    ws.cell(1, 4).value = "Fin garantie"
    ws.cell(1, 5).value = "Réforme"
    for imma in contrat_lnks.keys():
        if imma not in sn_imma.values():
            ws.cell(row_idx, 1).value = str(imma)
            ws.cell(row_idx, 2).value = all_eqpt_dict[imma]['n_seri']
            ws.cell(row_idx, 3).value = ', '.join(str(y[0]) + ': {:2.2f} €'.format(float(y[1])) for y in contrat_lnks[imma])
            ws.cell(row_idx, 4).value = all_eqpt_dict[imma]['fdg']
            ws.cell(row_idx, 5).value = all_eqpt_dict[imma]['date_refor']
            row_idx += 1
    p, n = os.path.split(FILENAME)
    wb.save(os.path.join(p, 'Xref-' + n))
